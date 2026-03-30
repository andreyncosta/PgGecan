"""
Load templates/gecan_schema_template.xlsx and fill the «unidades» sheet with N random agencies.
Writes: templates/gecan_100_agencias_aleatorias.xlsx

Run: python scripts/fill_template_random_agencies.py
"""

from __future__ import annotations

import random
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
TEMPLATE = ROOT / "templates" / "gecan_schema_template.xlsx"
OUT = ROOT / "templates" / "gecan_100_agencias_aleatorias.xlsx"

UFS = [
    "AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO", "MA", "MT", "MS", "MG",
    "PA", "PB", "PR", "PE", "PI", "RJ", "RN", "RS", "RO", "RR", "SC", "SP", "SE", "TO",
]

LOCALS = [
    "Centro", "Norte", "Sul", "Leste", "Oeste", "Jardim América", "Setor Bancário",
    "Asa Norte", "Asa Sul", "Taguatinga", "Ceilândia", "Samambaia", "Brazlândia",
    "Campo Grande", "Savassi", "Boa Viagem", "Paulista", "Copacabana", "Barra",
    "Pelotas", "Caxias", "Maringá", "Londrina", "Joinville", "Florianópolis",
]

TIPOS = ["Agência", "PA", "Escritório de Negócios"]

SUPERS = {
    "DF": "SUPVA DF", "SP": "SUPVA SP", "RJ": "SUPVA RJ", "MG": "SUPVA MG",
    "BA": "SUPVA BA", "RS": "SUPVA RS", "PR": "SUPVA PR", "PE": "SUPVA PE",
    "CE": "SUPVA CE", "GO": "SUPVA GO", "MS": "SUPVA MS",
}


def _rnd_money(rng: random.Random, lo: float, hi: float) -> float:
    return round(rng.uniform(lo, hi), 2)


def _rnd_int(rng: random.Random, lo: int, hi: int) -> int:
    return rng.randint(lo, hi)


def build_row(rng: random.Random, row_id: int) -> list:
    uf = rng.choice(UFS)
    local = rng.choice(LOCALS)
    tipo = rng.choices(TIPOS, weights=[0.72, 0.22, 0.06], k=1)[0]
    if tipo == "PA":
        nome = f"PA {local} {uf}"
    elif tipo == "Escritório de Negócios":
        nome = f"ESCRITÓRIO DE NEGÓCIOS {local}"
    else:
        nome = f"Agência {local} — {uf}"

    ra = f"{local.upper()} — RA {rng.randint(1, 30)}"
    uf_ra = f"{uf} {ra}"
    codigo = f"{rng.randint(1, 9999):04d}"
    clientes = _rnd_int(rng, 200, 45000)
    gerentes = max(1, min(12, clientes // 8000 + rng.randint(0, 3)))

    producao = _rnd_money(rng, 2e6, 120e6)
    resultado = _rnd_money(rng, -8e6, 12e6)
    rla = _rnd_money(rng, -2e6, 25e6)
    desp_adm = _rnd_money(rng, -2e6, -50_000)
    alugueis = rng.choice([0, _rnd_money(rng, -1.5e6, -20_000)])
    atend = _rnd_int(rng, 0, 45_000)
    caixa = _rnd_int(rng, 0, min(atend, 25_000))
    gerencia = _rnd_int(rng, 0, 5000)
    atend_col = max(0, atend - rng.randint(0, min(atend, 5000)))

    dea = round(rng.uniform(1.5, 18.0), 2)
    receita_colab = _rnd_money(rng, 0, 12e6)
    custo_at = round(rng.uniform(5, 4500), 2) if atend > 0 else 0.0
    pct_caixa = round(rng.uniform(0, 100), 1) if atend > 0 else 0.0

    metas = round(rng.uniform(65, 115), 1)
    m1s24 = round(rng.uniform(50, 115), 1)
    m2s24 = round(rng.uniform(50, 115), 1)
    m1s25 = round(rng.uniform(50, 115), 1)
    m2s25 = round(rng.uniform(50, 115), 1)

    ano = str(rng.randint(1965, 2025))
    alta_ma = _rnd_int(rng, 0, 8000)
    media_alta_ma = _rnd_int(rng, 0, 5000)
    media_ma = _rnd_int(rng, 0, 12000)
    baixa_ma = _rnd_int(rng, 0, 45000)
    middle = _rnd_int(rng, 0, 20)

    empresarial = _rnd_int(rng, 0, 120)
    empresas = _rnd_int(rng, 0, 300)
    empreendedor = _rnd_int(rng, 0, 800)

    consignado = _rnd_money(rng, 0, 45e6)
    imobiliario = _rnd_money(rng, 0, 80e6)
    parcelado_pf = _rnd_money(rng, 0, 12e6)
    parcelado_pj = _rnd_money(rng, 0, 15e6)
    plano_emp = _rnd_money(rng, 0, 0)
    rot_pf = _rnd_money(rng, 0, 3e6)
    rot_pj = _rnd_money(rng, 0, 2.5e6)
    agro = 0.0

    taxa_conv = str(round(rng.uniform(0, 25), rng.choice([2, 4, 8])))
    novo_mod = rng.choice(["Sim", "Não"])
    encerr = rng.choice(["NÃO", "NÃO", "NÃO", "SIM", ""])

    qt_contr = _rnd_int(rng, 50, 6500)
    vl_prod_vis = _rnd_money(rng, 0, 12e6)
    nr_port = _rnd_int(rng, 0, 5000)
    vl_seg = _rnd_money(rng, 0, 2e9)

    tipologia = "PA" if tipo == "PA" else ("Escritório de Negócios" if "ESCRITÓRIO" in nome else "Agência")
    endereco = f"{ra}, {uf}"
    metragem = _rnd_int(rng, 80, 1200)
    superint = SUPERS.get(uf, f"SUPVA {uf}")
    porte = min(5, max(1, clientes // 5000 + 1))
    horario = rng.choice(["08:00 às 15:00", "09:00 às 15:00", "10:00 às 16:00", "08:30 às 15:30"])
    dentro = 1 if rng.random() < 0.08 else 0

    total_func = gerentes + max(1, clientes // 800)
    escr = max(0, int(total_func * rng.uniform(0.25, 0.42)))
    cx_banc = max(0, int(total_func * rng.uniform(0.15, 0.28)))
    g_neg = max(0, gerentes - 1)
    g_exp = 1 if gerentes > 2 else 0
    g_geral = 1
    demais = max(0, total_func - escr - cx_banc - g_neg - g_exp - g_geral)
    qt_vis = int(qt_contr * rng.uniform(0.03, 0.12)) if vl_prod_vis > 0 else 0

    return [
        row_id,
        nome,
        codigo,
        uf,
        ra,
        uf_ra,
        resultado,
        _rnd_money(rng, -10e6, 8e6),
        rla,
        desp_adm,
        alugueis,
        round(rng.uniform(0, 25), 1),
        producao,
        atend,
        caixa,
        gerencia,
        atend_col,
        dea,
        clientes,
        gerentes,
        receita_colab,
        custo_at,
        pct_caixa,
        metas,
        m1s24,
        m2s24,
        m1s25,
        m2s25,
        ano,
        alta_ma,
        media_alta_ma,
        media_ma,
        baixa_ma,
        middle,
        empresarial,
        empresas,
        empreendedor,
        consignado,
        imobiliario,
        parcelado_pf,
        parcelado_pj,
        plano_emp,
        rot_pf,
        rot_pj,
        agro,
        taxa_conv,
        novo_mod,
        encerr,
        qt_contr,
        vl_prod_vis,
        nr_port,
        vl_seg,
        tipologia,
        endereco,
        metragem,
        superint,
        porte,
        horario,
        dentro,
        escr,
        g_neg,
        g_exp,
        g_geral,
        cx_banc,
        demais,
        qt_vis,
    ]


def main() -> int:
    if not TEMPLATE.is_file():
        print(f"Missing template: {TEMPLATE} — run generate_excel_template.py first", flush=True)
        return 1

    rng = random.Random(42)
    wb = load_workbook(TEMPLATE)
    ws = wb["unidades"]

    n = 100
    for i in range(n):
        row_values = build_row(rng, i + 1)
        r = 3 + i
        for col, val in enumerate(row_values, start=1):
            ws.cell(row=r, column=col, value=val)

    # meta: sample rows
    if "meta" in wb.sheetnames:
        wm = wb["meta"]
        wm["A3"] = "sample_seed"
        wm["B3"] = "random_100_v1"
        wm["A4"] = "gerado_em"
        wm["B4"] = datetime.now().isoformat(timespec="seconds")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"OK: {n} agencias -> {OUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
