"""
Generate Excel templates matching db/schema.sql (one sheet per table).
Run from project root: python scripts/generate_excel_template.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "templates" / "gecan_schema_template.xlsx"

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TYPE_FONT = Font(italic=True, color="64748B")
WRAP = Alignment(wrap_text=True, vertical="top")


def main() -> int:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # --- unidades ---
    ws = wb.active
    ws.title = "unidades"

    columns: list[tuple[str, str]] = [
        ("id", "INTEGER PRIMARY KEY — identificador único"),
        ("nome", "TEXT NOT NULL — nome da unidade"),
        ("codigo", "TEXT — código"),
        ("uf", "TEXT — UF"),
        ("ra", "TEXT — região administrativa"),
        ("uf_ra", "TEXT — UF + RA combinado"),
        ("resultado", "REAL"),
        ("resultado_com_dg", "REAL"),
        ("rla", "REAL"),
        ("desp_adm_total", "REAL"),
        ("alugueis", "REAL"),
        ("custo_ocupacao", "REAL"),
        ("producao_total", "REAL"),
        ("atendimento_total", "INTEGER"),
        ("caixa_total", "INTEGER"),
        ("gerencia_total", "INTEGER"),
        ("atendimento_col", "INTEGER"),
        ("dea", "REAL"),
        ("clientes", "INTEGER"),
        ("gerentes", "INTEGER"),
        ("receita_colab", "REAL"),
        ("custo_atendimento", "REAL"),
        ("pct_caixa", "REAL"),
        ("metas_media", "REAL"),
        ("metas_1s24", "REAL"),
        ("metas_2s24", "REAL"),
        ("metas_1s25", "REAL"),
        ("metas_2s25", "REAL"),
        ("data_abertura", "TEXT — ex.: ano ou data"),
        ("alta_ma", "INTEGER"),
        ("media_alta_ma", "INTEGER"),
        ("media_ma", "INTEGER"),
        ("baixa_ma", "INTEGER"),
        ("middle", "INTEGER"),
        ("empresarial", "INTEGER"),
        ("empresas", "INTEGER"),
        ("empreendedor", "INTEGER"),
        ("consignado", "REAL"),
        ("imobiliario", "REAL"),
        ("parcelado_pf", "REAL"),
        ("parcelado_pj", "REAL"),
        ("plano_empresario", "REAL"),
        ("rotativo_pf", "REAL"),
        ("rotativo_pj", "REAL"),
        ("agro", "REAL"),
        ("taxa_conversao", "TEXT"),
        ("novo_modelo", "TEXT"),
        ("encerramento", "TEXT"),
        ("qt_contratos", "INTEGER"),
        ("vl_producao_visita", "REAL"),
        ("nr_portabilidade", "INTEGER"),
        ("vl_seguros_total", "REAL"),
        ("tipologia", "TEXT"),
        ("endereco", "TEXT"),
        ("metragem_quadrada", "INTEGER"),
        ("superintendencia", "TEXT"),
        ("porte", "INTEGER"),
        ("horario_funcionamento", "TEXT"),
        ("dentro_orgao", "INTEGER NOT NULL DEFAULT 0 — use 0 ou 1 (booleano)"),
        ("escriturarios", "INTEGER"),
        ("gerente_negocio", "INTEGER"),
        ("gerente_expediente", "INTEGER"),
        ("gerente_geral", "INTEGER"),
        ("caixa_bancario", "INTEGER"),
        ("demais_cargos", "INTEGER"),
        ("qt_contratos_visita", "INTEGER"),
    ]

    for col, (name, hint) in enumerate(columns, start=1):
        c1 = ws.cell(row=1, column=col, value=name)
        c1.fill = HEADER_FILL
        c1.font = HEADER_FONT
        c1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c2 = ws.cell(row=2, column=col, value=hint)
        c2.font = TYPE_FONT
        c2.alignment = WRAP
        ws.column_dimensions[get_column_letter(col)].width = min(18, max(12, len(name) + 2))

    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 36

    # --- meta ---
    wm = wb.create_sheet("meta")
    wm.append(["key", "value"])
    wm.append(["TEXT PRIMARY KEY", "TEXT"])
    for cell in wm[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for cell in wm[2]:
        cell.font = TYPE_FONT
    wm.freeze_panes = "A3"
    wm.column_dimensions["A"].width = 28
    wm.column_dimensions["B"].width = 48

    # --- README (última aba) ---
    wr = wb.create_sheet("_schema_info")
    wr["A1"] = "GECAN — template Excel (espelha db/schema.sql)"
    wr["A1"].font = Font(bold=True, size=14)
    wr["A3"] = (
        "Preencha as abas «unidades» e «meta» conforme as colunas de gecan.db. "
        "Linha 1 = nomes das colunas SQLite; linha 2 = tipo / observação. "
        "Importação para SQLite pode ser feita via script Python ou export CSV a partir deste arquivo."
    )
    wr["A3"].alignment = Alignment(wrap_text=True)
    wr.column_dimensions["A"].width = 96

    wb.save(OUT)
    print(f"OK: {OUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
